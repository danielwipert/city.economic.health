#!/usr/bin/env python3
"""
CALCULATE ECONOMIC METRICS - RECONCILED VERSION
================================================
Combines the best of both approaches:
- Original: Solid data reading, clean architecture
- New: 3-component Cost of Living, Percentile-based scoring

Why this works:
- Reads from processed_economic_data_v2.json (correct data structure)
- Uses percentile scoring (handles outliers better than z-scores)
- Implements 3-component COL (affordability + direction + volatility)
- Maintains the original's reliable metric extraction
"""

import json
from pathlib import Path
from statistics import mean, stdev
from datetime import datetime, date
from typing import Dict, List, Optional

SCRIPT_DIR = Path(__file__).parent


# ============================================================================
# STEP 1: CALCULATE 2-COMPONENT OFFICE WORKER RATIO METRIC
# ============================================================================

def calculate_owr_final_score(metro_data: Dict, all_metros: List[Dict]) -> float:
    """
    Calculate office worker ratio score (0-5 points):
    - Component 1 (YoY change using 3-month avg): 0-3 points
      Going UP = better (more office workers)
    - Component 2 (absolute share of total employment): 0-2 points
      Higher % = better. Uses office_workers / all_employees (both from FRED LAUS),
      not office_workers / civilian_population — avoids stale population denominator
      and keeps both series on the same geographic footprint.
    Total: 0-5 points (higher is better)
    """
    office_workers_metric = metro_data['data'].get('office_workers')
    if not office_workers_metric:
        return None

    # Component 1: YoY change (3-month average)
    yoy_3mo = office_workers_metric.get('3month_avg_yoy')
    yoy_pct_change = yoy_3mo.get('pct_change') if yoy_3mo else None

    yoy_changes = []
    for metro in all_metros:
        ow = metro['data'].get('office_workers')
        if ow:
            y = ow.get('3month_avg_yoy')
            if y:
                pct = y.get('pct_change')
                if pct is not None:
                    yoy_changes.append(pct)

    if yoy_pct_change is not None and yoy_changes:
        better_count = sum(1 for v in yoy_changes if v < yoy_pct_change)
        c1_score = (better_count / len(yoy_changes)) * 3.0
    else:
        c1_score = 1.5

    # Component 2: office workers as share of total employment (not population)
    ow_latest = office_workers_metric.get('latest_value')
    emp_latest = metro_data['data'].get('all_employees', {}).get('latest_value')
    abs_pct = (ow_latest / emp_latest * 100) if ow_latest and emp_latest else None

    abs_pcts = []
    for metro in all_metros:
        ow = metro['data'].get('office_workers', {}).get('latest_value')
        emp = metro['data'].get('all_employees', {}).get('latest_value')
        if ow and emp:
            abs_pcts.append(ow / emp * 100)

    if abs_pct is not None and abs_pcts:
        better_count = sum(1 for v in abs_pcts if v < abs_pct)
        c2_score = (better_count / len(abs_pcts)) * 2.0
    else:
        c2_score = 1.0

    return min(5.0, max(0.0, c1_score + c2_score))


# ============================================================================
# STEP 1B: CALCULATE 3-COMPONENT COST OF LIVING METRIC
# ============================================================================

PSF_STALENESS_THRESHOLD_MONTHS = 9  # If PSF date lags earnings date by more than this, treat as missing


def _psf_is_stale(metro_data: Dict) -> bool:
    """
    Returns True if PSF data is more than PSF_STALENESS_THRESHOLD_MONTHS behind
    the earnings date. Stale PSF produces misleading YoY trend signals.
    """
    psf_date_str = metro_data['data'].get('price_per_sqft', {}).get('latest_date')
    earn_date_str = metro_data['data'].get('hourly_earnings', {}).get('latest_date')
    if not psf_date_str or not earn_date_str:
        return False
    psf_date = datetime.strptime(psf_date_str, '%Y-%m-%d').date()
    earn_date = datetime.strptime(earn_date_str, '%Y-%m-%d').date()
    months_behind = (earn_date.year - psf_date.year) * 12 + (earn_date.month - psf_date.month)
    return months_behind > PSF_STALENESS_THRESHOLD_MONTHS


def calculate_col_component1(metro_data: Dict) -> float:
    """
    Component 1: Absolute Affordability
    Price per square foot ÷ Hourly earnings (higher = worse affordability)
    Returns None if PSF data is stale (> 9 months behind earnings) to avoid
    using outdated housing prices in the ratio.
    """
    if 'price_per_sqft' not in metro_data['data']:
        return None
    if 'hourly_earnings' not in metro_data['data']:
        return None

    if _psf_is_stale(metro_data):
        return None

    psf = metro_data['data']['price_per_sqft'].get('latest_value')
    earnings = metro_data['data']['hourly_earnings'].get('latest_value')

    if psf is None or earnings is None or earnings == 0:
        return None

    col_index = psf / earnings
    return col_index


def calculate_col_component2(metro_data: Dict) -> float:
    """
    Component 2: Direction of Affordability — YoY % change in COL ratio.
    Returns the percentage change in (PSF / hourly_earnings) over 12 months.
    Positive = affordability worsening; negative = affordability improving.
    Returns 0 if data is unavailable.
    """
    if 'price_per_sqft' not in metro_data['data'] or 'hourly_earnings' not in metro_data['data']:
        return 0

    if _psf_is_stale(metro_data):
        return 0

    psf_metric      = metro_data['data']['price_per_sqft']
    earnings_metric = metro_data['data']['hourly_earnings']

    psf_current      = psf_metric.get('latest_value')
    earnings_current = earnings_metric.get('latest_value')

    if psf_current is None or earnings_current is None or earnings_current == 0:
        return 0

    col_current = psf_current / earnings_current

    psf_yoy      = psf_metric.get('yoy_change')
    earnings_yoy = earnings_metric.get('yoy_change')

    if psf_yoy and earnings_yoy:
        psf_prev      = psf_yoy.get('current') - psf_yoy.get('change')
        earnings_prev = earnings_yoy.get('current') - earnings_yoy.get('change')

        if psf_prev and earnings_prev and earnings_prev != 0:
            col_prev = psf_prev / earnings_prev
            if col_prev != 0:
                col_yoy_pct = (col_current - col_prev) / col_prev * 100
                return col_yoy_pct

    return 0


def calculate_col_component3(metro_data: Dict) -> float:
    """
    Component 3: Peer-relative affordability trend.
    Returns the same YoY % change in COL ratio as component 2.
    The actual peer-relative score is computed inside calculate_col_final_score,
    where the full cross-city distribution is available to derive the median.
    """
    return calculate_col_component2(metro_data)


def calculate_col_final_score(metro_data: Dict, all_metros: List[Dict], 
                             col_component2_all: Dict, col_component3_all: Dict) -> float:
    """
    Combine 3 COL components into final 0-10 score (lower = better affordability):
    - Component 1 (absolute affordability): 0-5 pts — min-max normalized PSF/earnings ratio.
        Dominates at 50% so expensive cities can't escape their absolute cost via trend credit alone.
    - Component 2 (direction, graduated): 0-3 pts — YoY % change in COL ratio;
        ≤ -5% (improving) → 0, flat → 1.5, ≥ +5% (worsening) → 3
    - Component 3 (peer-relative trend): 0-2 pts — deviation from national median YoY%;
        improving much faster than peers → 0, in line → 1, worsening much faster → 2
    """
    metro_name = metro_data['metro_name']
    
    # Component 1: Absolute affordability (0-3 points)
    col_index = calculate_col_component1(metro_data)
    if col_index is None:
        return None
    
    # Find where this metro ranks in affordability
    col_indices = []
    for metro in all_metros:
        idx = calculate_col_component1(metro)
        if idx is not None:
            col_indices.append(idx)
    
    if not col_indices:
        return None
    
    # Normalize to 0-3 scale (higher is worse affordability)
    min_col = min(col_indices)
    max_col = max(col_indices)
    
    if max_col > min_col:
        c1_score = ((col_index - min_col) / (max_col - min_col)) * 5.0
    else:
        c1_score = 2.5
    
    # Component 2: Direction of change (0-4 points) — graduated, not binary.
    # Uses YoY % change in COL ratio. Neutral at 0%, full penalty at ±5%.
    # Old binary cliff (any positive → 4.0) wrongly gave max penalty to cities
    # with tiny upticks while still being far more affordable than coastal metros.
    col_yoy_pct = col_component2_all.get(metro_name, 0)
    if col_yoy_pct <= -5:
        c2_score = 0.0   # Strong improvement → lowest penalty
    elif col_yoy_pct >= 5:
        c2_score = 3.0   # Strong worsening  → highest penalty
    else:
        c2_score = (col_yoy_pct + 5) / 10.0 * 3.0  # Linear -5%→0→+5% maps to 0→1.5→3

    # Component 3: Peer-relative trend (0-3 points).
    # Measures how this city's affordability trend compares to the national median.
    # A city worsening much faster than peers is penalized; improving faster is rewarded.
    # This replaces the old "volatility" proxy (avg of |PSF_YoY%| + |earnings_YoY%|)
    # which double-counted c2 and wrongly penalized all cities with any price movement.
    all_yoy_pcts = [v for v in col_component3_all.values() if v is not None]
    if all_yoy_pcts:
        sorted_pcts = sorted(all_yoy_pcts)
        n = len(sorted_pcts)
        median_yoy = (sorted_pcts[n // 2 - 1] + sorted_pcts[n // 2]) / 2 if n % 2 == 0 else sorted_pcts[n // 2]
    else:
        median_yoy = 0

    deviation = col_yoy_pct - median_yoy  # positive = worsening vs peers
    if deviation <= -10:
        c3_score = 0.0   # Improving much faster than peers → lowest penalty
    elif deviation >= 10:
        c3_score = 2.0   # Worsening much faster than peers → highest penalty
    else:
        c3_score = (deviation + 10) / 20.0 * 2.0  # Linear ±10% range maps to 0→1→2
    
    total = c1_score + c2_score + c3_score
    return min(10.0, max(0.0, total))


# ============================================================================
# 101A: UNEMPLOYMENT 2-COMPONENT COMPOSITE
# ============================================================================

def calculate_unemployment_composite(unemp_level, unemp_yoy_pp):
    """
    2-component composite for Unemployment Rate (101A). Total: 0-10, higher = better.
    Percentile-ranked with invert=False.

    c1 (75% = 7.5 pts): Current unemployment level.
        Absolute scale anchored at 2% (full employment floor) and 8% (distress ceiling).
        Lower unemployment = higher score.
        2% or below -> 7.5 pts | 5% -> 3.75 pts | 8% or above -> 0 pts.

    c2 (25% = 2.5 pts): YoY direction (absolute pp change, latest vs same month prior year).
        Improvement rewarded, deterioration penalized. Neutral at 0pp change.
        Uses point-to-point YoY (not 3-month avg) — LAUS data has a structural
        publication gap at index 12 for ~46/50 metros each run; point-to-point
        avoids nulls. A 13-month fallback handles the rare case where the
        year-ago month itself was never published by BLS.
        -1.0pp or better (improving) -> 2.5 pts | flat -> 1.25 pts | +1.0pp or worse -> 0 pts.

    Why not 3-month average YoY: BLS LAUS releases create a systematic gap at the
    third-most-recent month (Oct 2025 missing for 46/50 metros as of April 2026),
    which would null out the 3-month average for almost the entire dataset each run.
    """
    EXCELLENT = 2.0   # at or below: maximum c1 score
    DISTRESS  = 8.0   # at or above: zero c1 score

    # c1: level component (0-7.5 pts)
    if unemp_level is None:
        c1 = 3.75  # neutral default
    else:
        c1 = max(0.0, min(7.5, (DISTRESS - unemp_level) / (DISTRESS - EXCELLENT) * 7.5))

    # c2: direction component (0-2.5 pts)
    # unemp_yoy_pp is the absolute pp change (positive = deteriorating)
    RANGE = 1.0  # ±1.0pp captures meaningful movement without overreacting to noise
    if unemp_yoy_pp is None:
        c2 = 1.25  # neutral default
    else:
        c2 = max(0.0, min(2.5, (RANGE - unemp_yoy_pp) / (2 * RANGE) * 2.5))

    return min(10.0, max(0.0, c1 + c2))


# ============================================================================
# 204A: DAYS ON MARKET 2-COMPONENT COMPOSITE
# ============================================================================

def calculate_dom_composite_score(days_yoy, dom_level):
    """
    2-component composite for Days on Market (204A). Total: 0-10, higher = better.
    Percentile-ranked with invert=False.

    c1 (60% = 6 pts): YoY % change in DoM — direction interpretation blends across a
        transition zone (45–75 days) rather than switching at a hard 60-day cliff.

        Below 45 days (tight market): loosening = good for incoming workers.
            -30% or worse → 0 pts | flat → 3 pts | +30% or more → 6 pts.
        Above 75 days (elevated/soft market): loosening = demand destruction = bad.
            Direction inverts: +30% or worse → 0 pts | flat → 3 pts | -30% → 6 pts.
        45–75 days (transition zone): linearly blended between both interpretations.
            A market at 60 days (midpoint) weights each regime equally.

        Rationale: a market softening from 25→40 days gains healthy inventory.
        A market softening from 80→95 days signals buyers cannot transact —
        labor mobility is impaired. The blend zone eliminates the cliff where a
        1-day difference in DoM level caused a multi-point score swing.

    c2 (40% = 4 pts): Absolute DoM level — context and distress filter.
        Normal range (35-80 days) peaks at 4 pts.
        Extreme tightness (<15 days) → 0 pts — accessibility problem.
        Severe softness (>130 days) → 0 pts — potential demand destruction.
    """
    BLEND_LOW  = 45   # below this: fully in "loosening = good" regime
    BLEND_HIGH = 75   # above this: fully in "loosening = bad" regime

    # --- c1: Trend (0-6 points) — blended regime ---
    if days_yoy is None:
        c1_score = 3.0  # default to midpoint
    else:
        # Score under each pure regime (clamped to [0,6])
        c1_healthy  = max(0.0, min(6.0, (days_yoy + 30) / 60.0 * 6.0))   # loosening = good
        c1_elevated = max(0.0, min(6.0, (-days_yoy + 30) / 60.0 * 6.0))  # loosening = bad

        if dom_level is None or dom_level <= BLEND_LOW:
            c1_score = c1_healthy
        elif dom_level >= BLEND_HIGH:
            c1_score = c1_elevated
        else:
            # Linear blend: weight_elevated rises from 0 at BLEND_LOW to 1 at BLEND_HIGH
            w_elevated = (dom_level - BLEND_LOW) / (BLEND_HIGH - BLEND_LOW)
            c1_score = (1 - w_elevated) * c1_healthy + w_elevated * c1_elevated

    # --- c2: Level context (0-4 points) ---
    TIGHT_FLOOR  = 15   # at or below: full tightness penalty
    NORMAL_LOW   = 35   # lower bound of healthy/accessible range
    NORMAL_HIGH  = 80   # upper bound of healthy range
    SOFT_CEILING = 130  # at or above: full softness/distress penalty

    if dom_level is None:
        c2_score = 2.0  # default to midpoint
    elif NORMAL_LOW <= dom_level <= NORMAL_HIGH:
        c2_score = 4.0  # in the healthy zone
    elif dom_level < NORMAL_LOW:
        # Linear from 0 (at TIGHT_FLOOR) → 4 (at NORMAL_LOW)
        c2_score = max(0.0, (dom_level - TIGHT_FLOOR) / (NORMAL_LOW - TIGHT_FLOOR) * 4.0)
    else:
        # Linear from 4 (at NORMAL_HIGH) → 0 (at SOFT_CEILING)
        c2_score = max(0.0, 4.0 - (dom_level - NORMAL_HIGH) / (SOFT_CEILING - NORMAL_HIGH) * 4.0)

    return min(10.0, max(0.0, c1_score + c2_score))


# ============================================================================
# 107E/106D: LABOR DEMAND COMPOSITE
# ============================================================================

def calculate_labor_demand_composite(emp_yoy, wh_deviation):
    """
    Labor Demand Composite — replaces separate 107E and 106D metrics.
    0-10 scale, higher = stronger labor demand. Percentile-ranked invert=False.

    c1 (70% = 7 pts): Employment growth YoY — primary demand signal.
        Scale: -2% → 0 pts | 0% → 3.5 pts | +3% → 7 pts (linear).

    c2 (30% = 3 pts): Weekly hours deviation, employment-conditioned.
        The direction of the hours signal depends on whether employment is growing
        or contracting, capturing four economically distinct scenarios:

        STRONG  (emp ≥ 0, hours above trend): genuine demand confirmation — rewarded.
        GROWING (emp ≥ 0, hours below trend): jobs expanding but hours soft — mild caution.
        SQUEEZE (emp < 0, hours above trend): survivor squeeze — employers cutting
            headcount while squeezing remaining workers harder. Penalized.
        WEAK    (emp < 0, hours below trend): consistent contraction — neutral on c2,
            c1 already captures the employment decline.

    Combined weight: 25% (absorbs former 107E 15% + 106D 10%).
    """
    # c1: Employment growth (0-7 pts)
    if emp_yoy is None:
        c1 = 3.5  # neutral default
    else:
        c1 = max(0.0, min(7.0, (emp_yoy + 2.0) / 5.0 * 7.0))

    # c2: Hours deviation, employment-conditioned (0-3 pts)
    if wh_deviation is None:
        c2 = 1.5  # neutral default
    elif emp_yoy is None or emp_yoy >= 0:
        # Growing/stable employment: hours above trend confirms demand
        # Scale: -1.5% → 0 pts | 0% → 1.5 pts | +1.5% → 3 pts
        c2 = max(0.0, min(3.0, (wh_deviation + 1.5) / 3.0 * 3.0))
    else:
        # Contracting employment: hours above trend = survivor squeeze (penalized)
        # Hours below trend = consistent contraction (neutral — c1 already captures it)
        if wh_deviation >= 0:
            squeeze_penalty = min(1.5, wh_deviation / 1.5 * 1.5)
            c2 = max(0.0, 1.5 - squeeze_penalty)
        else:
            c2 = 1.5  # consistent contraction, no bonus/penalty on c2

    return min(10.0, max(0.0, c1 + c2))


# ============================================================================
# STEP 2: PERCENTILE-BASED SCORING
# ============================================================================

def calculate_percentile_score(value: float, all_values: List[float], 
                              invert: bool = False) -> float:
    """
    Calculate percentile score for a single value (0-100 scale)
    
    invert=False: Higher value = higher percentile (good for earnings, permits, etc)
    invert=True:  Lower value = higher percentile (good for unemployment, COL)
    """
    if not all_values or len(all_values) < 2:
        return 50.0
    
    valid_values = [v for v in all_values if v is not None]
    if not valid_values:
        return 50.0
    
    sorted_vals = sorted(valid_values)
    
    if invert:
        # For inverted metrics: lower is better
        # Count how many values are WORSE (higher) than this one
        better_count = sum(1 for v in sorted_vals if v > value)
    else:
        # For normal metrics: higher is better
        # Count how many values are WORSE (lower) than this one
        better_count = sum(1 for v in sorted_vals if v < value)
    
    percentile = (better_count / len(sorted_vals)) * 100.0
    return max(0, min(100, percentile))


# ============================================================================
# STEP 3: MAIN CALCULATION FUNCTION
# ============================================================================

def calculate_metrics():
    """Main calculation logic"""
    
    print("\n" + "=" * 80)
    print("CALCULATING ECONOMIC METRICS (RECONCILED VERSION)")
    print("=" * 80)
    
    # Load processed data
    print("\nSTEP 1: Loading processed economic data...")
    processed_path = SCRIPT_DIR / 'processed_economic_data_v2.json'
    
    if not processed_path.exists():
        print(f"❌ ERROR: {processed_path} not found")
        return None
    
    try:
        with open(processed_path, 'r', encoding='utf-8') as f:
            processed_data = json.load(f)
        print(f"✓ Loaded {len(processed_data['metros'])} metros")
    except Exception as e:
        print(f"❌ ERROR loading data: {e}")
        return None
    
    all_metros = list(processed_data['metros'].values())
    
    # Calculate COL components for all metros
    print("\nSTEP 2: Calculating 3-component Cost of Living...")
    
    col_component2_all = {}
    col_component3_all = {}
    
    for metro in all_metros:
        metro_name = metro['metro_name']
        col_component2_all[metro_name] = calculate_col_component2(metro)
        col_component3_all[metro_name] = calculate_col_component3(metro)
    
    print("✓ Calculated COL components")
    
    # Collect all metric values for percentile calculation
    print("\nSTEP 3: Collecting metrics for percentile ranking...")
    
    # Pre-calculate OWR scores for all metros
    owr_scores_all = {}
    for metro in all_metros:
        owr_score = calculate_owr_final_score(metro, all_metros)
        owr_scores_all[metro['metro_name']] = owr_score
    
    metrics_data = {
        '101A': [],  # Unemployment rate
        '102A': [],  # Civilian labor force YoY % change
        '103B': [],  # Hourly earnings YoY
        '104C': [],  # Cost of living (final score)
        '105C': [],  # Office worker ratio (2-component score)
        '107E': [],  # Labor demand composite (employment growth + hours, employment-conditioned)
        '200B': [],  # Building permits YoY
        '204A': [],  # Median days on market
    }

    for metro in all_metros:
        data = metro['data']

        # Extract latest values for each metric
        unemp       = data.get('unemployment_rate', {}).get('latest_value')
        unemp_yoy_pp = data.get('unemployment_rate', {}).get('yoy_change', {}).get('change')
        # 101A: 2-component composite — level (75%) + YoY direction (25%).
        unemp_composite = calculate_unemployment_composite(unemp, unemp_yoy_pp)
        # 102A: YoY % change in civilian labor force — avoids denominator/population
        # mismatch issues with LFP rate (stale Census pop + BLS vs ACS geo mismatch).
        # Captures supply-side momentum: labor force growing = workers moving in or
        # re-entering. Differentiates from 107E (employment demand) by measuring
        # availability, not hiring activity.
        clf_yoy = data.get('civilian_labor_force', {}).get('yoy_change', {}).get('pct_change')
        # 103B: 3-month average YoY — smoother than point-to-point; earnings data
        # has no missing-value gaps across the 50-metro dataset so the 3mo avg is
        # always available. Avoids single-month noise (year-end bonus mix, seasonal
        # staffing) that causes >1pp swings vs the underlying trend for ~40% of metros.
        earnings_yoy = data.get('hourly_earnings', {}).get('3month_avg_yoy', {}).get('pct_change')
        col_score = calculate_col_final_score(metro, all_metros, col_component2_all, col_component3_all)
        owr = owr_scores_all.get(metro['metro_name'])
        wh_data = data.get('weekly_hours', {})
        wh_3mo = wh_data.get('3month_average')
        wh_12mo = wh_data.get('12month_average')
        wh_deviation = ((wh_3mo - wh_12mo) / wh_12mo * 100
                        if wh_3mo is not None and wh_12mo and wh_12mo != 0
                        else None)
        emp_yoy = data.get('all_employees', {}).get('yoy_change', {}).get('pct_change')
        # 107E: Labor Demand Composite — employment growth (primary) + hours deviation
        # (employment-conditioned). Combines former 107E and 106D into one signal.
        ldc = calculate_labor_demand_composite(emp_yoy, wh_deviation)
        # 200B: Staleness check — monthly series: 6-month threshold.
        # Annual series (e.g. Fresno county): 18-month threshold so the most
        # recent annual total is still usable. For annual series, fall back to
        # point-to-point yoy_change since 3month_avg_yoy is intentionally None.
        _bp = data.get('building_permits', {})
        _bp_date_str = _bp.get('latest_date')
        _bp_is_annual = bool(_bp.get('yoy_change', {}).get('is_annual'))
        _bp_stale = False
        if _bp_date_str:
            _bp_date = datetime.strptime(_bp_date_str, '%Y-%m-%d').date()
            _bp_threshold = 1095 if _bp_is_annual else 180  # annual: 3yr; monthly: 6mo
            _bp_stale = (date.today() - _bp_date).days > _bp_threshold
        if _bp_stale:
            permits_yoy = None
        elif _bp_is_annual:
            permits_yoy = _bp.get('yoy_change', {}).get('pct_change')
        else:
            permits_yoy = _bp.get('3month_avg_yoy', {}).get('pct_change')
        # 204A: 2-component composite — trend (60%) + level context (40%).
        days_yoy   = data.get('median_days_on_market', {}).get('yoy_change', {}).get('pct_change')
        dom_level  = data.get('median_days_on_market', {}).get('latest_value')
        dom_composite = calculate_dom_composite_score(days_yoy, dom_level)

        # Add to collections (only if not None)
        if unemp is not None:
            metrics_data['101A'].append(unemp_composite)
        if clf_yoy is not None:
            metrics_data['102A'].append(clf_yoy)
        if earnings_yoy is not None:
            metrics_data['103B'].append(earnings_yoy)
        if col_score is not None:
            metrics_data['104C'].append(col_score)
        if owr is not None:
            metrics_data['105C'].append(owr)
        if emp_yoy is not None or wh_deviation is not None:
            metrics_data['107E'].append(ldc)
        if permits_yoy is not None:
            metrics_data['200B'].append(permits_yoy)
        if days_yoy is not None or dom_level is not None:
            metrics_data['204A'].append(dom_composite)
    
    print(f"✓ Collected metrics from {len(all_metros)} metros")
    
    # Weight configuration — total = 100
    # Employment (85%): unemployment, LFP, earnings, COL, OWR, labor demand composite
    # Housing (15%): building permits, days on market
    # 107E is now the Labor Demand Composite (employment growth + employment-conditioned hours).
    # Absorbs former 107E (15%) + 106D (10%) = 25% combined.
    weights = {
        '101A': 20,  # Unemployment (lower is better)
        '102A': 10,  # Civilian labor force YoY % change (higher is better) — supply-side momentum
        '103B': 15,  # Earnings growth YoY (higher is better) — increased from 10%; real-time demand signal
        '104C': 12,  # Cost of living (lower is better) — increased from 10%; key business location signal
        '105C': 3,   # Office worker ratio (higher is better) — reduced from 5%; tiebreaker signal only
        '107E': 25,  # Labor demand composite: employment growth + hours (employment-conditioned)
        '200B': 10,  # Building permits YoY (higher is better)
        '204A': 5,   # Days on market composite
    }
    
    # Calculate percentile scores for each metro
    print("\nSTEP 4: Calculating percentile scores...")
    
    results = []
    
    for i, metro in enumerate(all_metros):
        metro_name = metro['metro_name']
        rank = metro.get('rank', i + 1)
        primary_city = metro.get('primary_city', '')
        data = metro['data']
        
        # Extract values
        unemp        = data.get('unemployment_rate', {}).get('latest_value')
        unemp_yoy_pp = data.get('unemployment_rate', {}).get('yoy_change', {}).get('change')
        unemp_composite = calculate_unemployment_composite(unemp, unemp_yoy_pp)
        clf_yoy = data.get('civilian_labor_force', {}).get('yoy_change', {}).get('pct_change')
        earnings_yoy = data.get('hourly_earnings', {}).get('3month_avg_yoy', {}).get('pct_change')
        col_score = calculate_col_final_score(metro, all_metros, col_component2_all, col_component3_all)
        owr = owr_scores_all.get(metro_name)
        wh_data = data.get('weekly_hours', {})
        wh_3mo = wh_data.get('3month_average')
        wh_12mo = wh_data.get('12month_average')
        wh_deviation = ((wh_3mo - wh_12mo) / wh_12mo * 100
                        if wh_3mo is not None and wh_12mo and wh_12mo != 0
                        else None)
        emp_yoy      = data.get('all_employees', {}).get('yoy_change', {}).get('pct_change')
        ldc          = calculate_labor_demand_composite(emp_yoy, wh_deviation)
        _bp2 = data.get('building_permits', {})
        _bp_date_str2 = _bp2.get('latest_date')
        _bp_is_annual2 = bool(_bp2.get('yoy_change', {}).get('is_annual'))
        _bp_stale2 = False
        if _bp_date_str2:
            _bp_date2 = datetime.strptime(_bp_date_str2, '%Y-%m-%d').date()
            _bp_threshold2 = 1095 if _bp_is_annual2 else 180  # annual: 3yr; monthly: 6mo
            _bp_stale2 = (date.today() - _bp_date2).days > _bp_threshold2
        if _bp_stale2:
            permits_yoy = None
        elif _bp_is_annual2:
            permits_yoy = _bp2.get('yoy_change', {}).get('pct_change')
        else:
            permits_yoy = _bp2.get('3month_avg_yoy', {}).get('pct_change')
        days_yoy     = data.get('median_days_on_market', {}).get('yoy_change', {}).get('pct_change')
        dom_level    = data.get('median_days_on_market', {}).get('latest_value')
        dom_composite = calculate_dom_composite_score(days_yoy, dom_level)

        # Calculate percentile scores
        percentiles = {
            '101A': calculate_percentile_score(unemp_composite, metrics_data['101A'], invert=False),
            '102A': calculate_percentile_score(clf_yoy, metrics_data['102A'], invert=False) if clf_yoy is not None else 50,
            '103B': calculate_percentile_score(earnings_yoy, metrics_data['103B'], invert=False) if earnings_yoy is not None else 50,
            '104C': calculate_percentile_score(col_score, metrics_data['104C'], invert=True) if col_score else 50,
            '105C': calculate_percentile_score(owr, metrics_data['105C'], invert=False) if owr is not None else 50,
            '107E': calculate_percentile_score(ldc, metrics_data['107E'], invert=False),
            '200B': calculate_percentile_score(permits_yoy, metrics_data['200B'], invert=False) if permits_yoy is not None else 50,
            '204A': calculate_percentile_score(dom_composite, metrics_data['204A'], invert=False),
        }
        
        # Calculate weighted percentile score
        total_weight = sum(weights.values())
        weighted_percentile = sum(percentiles[code] * weights[code] for code in weights) / total_weight
        weighted_score = int(round(weighted_percentile))
        
        # Assign grade based on weighted percentile score.
        # Thresholds are calibrated to the achievable range (~24-70) that
        # results from averaging 10 percentile scores across 50 metros —
        # no city can realistically score above ~70 on the weighted average.
        if weighted_percentile >= 68:
            grade_letter, emoji, description = "A+", "🚀", "Excellent"
        elif weighted_percentile >= 63:
            grade_letter, emoji, description = "A", "✅", "Very Good"
        elif weighted_percentile >= 59:
            grade_letter, emoji, description = "A-", "👍", "Good"
        elif weighted_percentile >= 55:
            grade_letter, emoji, description = "B+", "📈", "Above Average"
        elif weighted_percentile >= 50:
            grade_letter, emoji, description = "B", "➡️", "Average"
        elif weighted_percentile >= 44:
            grade_letter, emoji, description = "B-", "⚠️", "Below Average"
        elif weighted_percentile >= 38:
            grade_letter, emoji, description = "C+", "📉", "Poor"
        elif weighted_percentile >= 32:
            grade_letter, emoji, description = "C", "⛔", "Very Poor"
        elif weighted_percentile >= 26:
            grade_letter, emoji, description = "C-", "🚨", "Critical"
        else:
            grade_letter, emoji, description = "D", "💥", "Emergency"
        
        results.append({
            "metro_name": metro_name,
            "rank": rank,
            "primary_city": primary_city,
            "raw_values": {
                "101A_composite": round(unemp_composite, 2),
                "101A_unemployment": round(unemp, 2) if unemp else None,
                "101A_unemp_yoy_pp": round(unemp_yoy_pp, 2) if unemp_yoy_pp is not None else None,
                "102A_clf_yoy": round(clf_yoy, 2) if clf_yoy is not None else None,
                "103B_earnings_yoy": round(earnings_yoy, 2) if earnings_yoy is not None else None,
                "104C_col": round(col_score, 2) if col_score else None,
                "105C_owr": round(owr, 2) if owr is not None else None,
                "107E_ldc_composite": round(ldc, 2),
                "107E_employment_growth_yoy": round(emp_yoy, 2) if emp_yoy is not None else None,
                "107E_wh_trend_deviation_pct": round(wh_deviation, 3) if wh_deviation is not None else None,
                "200B_permits_yoy": round(permits_yoy, 2) if permits_yoy is not None else None,
                "204A_dom_composite": round(dom_composite, 2),
                "204A_dom_yoy_pct": round(days_yoy, 2) if days_yoy is not None else None,
                "204A_dom_level_days": round(dom_level, 0) if dom_level is not None else None,
            },
            "percentile_scores": {code: round(p, 1) for code, p in percentiles.items()},
            "scores_100": {code: int(round(p)) for code, p in percentiles.items()},
            "weighted_percentile": round(weighted_percentile, 1),
            "weighted_score": weighted_score,
            "grade": {
                "letter": grade_letter,
                "emoji": emoji,
                "description": description,
                "percentile": round(weighted_percentile, 1)
            },
            "grade_display": f"{emoji} {grade_letter}"
        })
    
    print(f"✓ Calculated percentile scores for {len(results)} metros")
    
    # Create output
    output = {
        "calculation_timestamp": datetime.now().isoformat(),
        "calculation_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "version": "5.0",
        "scoring_method": "Percentile-Based (Rank Order)",
        "rubric": "9-Metric Percentile System: 85% Employment / 15% Housing",
        "note": "HPI and PSF YoY removed — housing price appreciation conflates supply constraints with genuine demand; payroll growth (107E) captures demand directly. 106D weekly hours scored as deviation from own 12-month trend to remove industry-composition bias. 104C COL: absolute affordability (3pts) + direction (4pts) + volatility (3pts). 105C OWR: YoY growth (3pts) + absolute % (2pts). 200B permits uses 3-month smoothed YoY.",
        "weight_configuration": weights,
        "score_codes": {
            "101A": "unemployment_rate (20)",
            "102A": "labor_force_participation (15)",
            "103B": "hourly_earnings_yoy (10)",
            "104C": "cost_of_living_3component (10)",
            "105C": "office_worker_ratio_2component (5)",
            "106D": "weekly_hours_trend_deviation_pct (10)",
            "107E": "total_nonfarm_employment_growth_yoy (15)",
            "200B": "building_permits_3month_smoothed_yoy (10)",
            "204A": "median_days_on_market_2component_composite (5)",
        },
        "metros": results
    }
    
    return output


# ============================================================================
# STEP 4: CREATE EXCEL FILE WITH RESULTS
# ============================================================================

def create_excel_from_metrics(output_data):
    """Create Excel file with all metrics organized in 5 sheets"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    print("\nSTEP 5: Creating Excel file...")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Color mapping for grades
    grade_colors = {
        'A+': 'C6EFCE', 'A': 'C6EFCE', 'A-': 'D4EDDA',
        'B+': 'D1E7DD', 'B': 'E2F0E1', 'B-': 'FFF2CC',
        'C+': 'FFEB9C', 'C': 'FFE699', 'C-': 'F8CBAD', 'D': 'F4B084'
    }
    
    def style_cell(cell, fill_color=None, bold=False, center=False):
        cell.border = border
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        if bold:
            cell.font = Font(bold=True)
        if center:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ========== SHEET 1: SUMMARY ==========
    ws_summary = wb.create_sheet("Summary", 0)
    
    headers = ['Rank', 'Metro Name', 'Primary City', 'Weighted Score', 'Grade', 'Percentile', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        style_cell(cell, center=True)
    
    sorted_metros = sorted(output_data['metros'], key=lambda x: x['weighted_score'], reverse=True)
    
    for row, metro in enumerate(sorted_metros, 2):
        ws_summary.cell(row=row, column=1, value=metro['rank'])
        ws_summary.cell(row=row, column=2, value=metro['metro_name'])
        ws_summary.cell(row=row, column=3, value=metro['primary_city'])
        ws_summary.cell(row=row, column=4, value=metro['weighted_score'])
        ws_summary.cell(row=row, column=5, value=metro['grade']['letter'])
        ws_summary.cell(row=row, column=6, value=metro['grade']['percentile'])
        ws_summary.cell(row=row, column=7, value=metro['grade']['description'])
        
        grade_color = grade_colors.get(metro['grade']['letter'], 'FFFFFF')
        for col in range(1, len(headers) + 1):
            cell = ws_summary.cell(row=row, column=col)
            style_cell(cell, fill_color=grade_color)
            if col == 4:
                cell.font = Font(bold=True)
            if col in [1, 4, 6]:
                cell.alignment = Alignment(horizontal='center')
    
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 10
    ws_summary.column_dimensions['F'].width = 12
    ws_summary.column_dimensions['G'].width = 15
    
    # ========== SHEET 2: METRIC SCORES ==========
    ws_scores = wb.create_sheet("Metric Scores", 1)
    
    metric_codes = ['101A', '102A', '103B', '104C', '105C', '106D', '107E', '200B', '204A']
    metric_names = {
        '101A': 'Unemployment', '102A': 'LFP', '103B': 'Earnings YoY', '104C': 'COL',
        '105C': 'Office Workers', '106D': 'Weekly Hours', '107E': 'Emp Growth YoY',
        '200B': 'Permits YoY', '204A': 'DoM YoY'
    }
    
    headers_scores = ['Metro Name', 'Weighted Score'] + [f"{code}\n({metric_names[code]})" for code in metric_codes]
    for col, header in enumerate(headers_scores, 1):
        cell = ws_scores.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        style_cell(cell, center=True)
    
    for row, metro in enumerate(sorted_metros, 2):
        ws_scores.cell(row=row, column=1, value=metro['metro_name'])
        ws_scores.cell(row=row, column=2, value=metro['weighted_score'])
        style_cell(ws_scores.cell(row=row, column=2), bold=True)
        
        for col, code in enumerate(metric_codes, 3):
            score = metro['scores_100'].get(code)
            cell = ws_scores.cell(row=row, column=col, value=score)
            style_cell(cell, center=True)
            if score:
                color = 'C6EFCE' if score >= 70 else ('FFF2CC' if score >= 50 else 'F8CBAD')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    ws_scores.column_dimensions['A'].width = 35
    ws_scores.column_dimensions['B'].width = 15
    for col in range(3, len(metric_codes) + 3):
        ws_scores.column_dimensions[get_column_letter(col)].width = 14
    
    # ========== SHEET 3: RAW VALUES ==========
    ws_raw = wb.create_sheet("Raw Values", 2)
    
    raw_metric_names = {
        '101A_unemployment': 'Unemployment %', '102A_clf_yoy': 'Labor Force YoY %',
        '103B_earnings_yoy': 'Earnings YoY %', '104C_col': 'COL Score',
        '105C_owr': 'Office Worker %', '106D_wh_trend_deviation_pct': 'WH Trend Dev %',
        '107E_employment_growth_yoy': 'Emp Growth YoY %', '200B_permits_yoy': 'Permits YoY %',
        '204A_dom_composite': 'DoM Composite (0-10)',
        '204A_dom_yoy_pct': 'DoM YoY %', '204A_dom_level_days': 'DoM Level (days)'
    }
    
    raw_keys = list(raw_metric_names.keys())
    headers_raw = ['Metro Name', 'Score/Grade'] + [raw_metric_names[k] for k in raw_keys]
    
    for col, header in enumerate(headers_raw, 1):
        cell = ws_raw.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        style_cell(cell, center=True)
    
    for row, metro in enumerate(sorted_metros, 2):
        ws_raw.cell(row=row, column=1, value=metro['metro_name'])
        score_grade = f"{metro['weighted_score']} ({metro['grade']['letter']})"
        cell = ws_raw.cell(row=row, column=2, value=score_grade)
        style_cell(cell, bold=True, center=True)
        
        for col, key in enumerate(raw_keys, 3):
            value = metro['raw_values'].get(key)
            cell = ws_raw.cell(row=row, column=col, value=value)
            style_cell(cell, center=True)
    
    ws_raw.column_dimensions['A'].width = 35
    ws_raw.column_dimensions['B'].width = 15
    for col in range(3, len(raw_keys) + 3):
        ws_raw.column_dimensions[get_column_letter(col)].width = 16
    
    # ========== SHEET 4: BY RANK ==========
    ws_rank = wb.create_sheet("By Rank", 3)
    
    by_rank = sorted(output_data['metros'], key=lambda x: x['rank'])
    headers_rank = ['Rank', 'Metro Name', 'Primary City', 'Weighted Score', 'Grade', 'Percentile', 
                    'Unemployment', 'LFP %', 'Earnings YoY', 'COL Score']
    
    for col, header in enumerate(headers_rank, 1):
        cell = ws_rank.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        style_cell(cell, center=True)
    
    for row, metro in enumerate(by_rank, 2):
        ws_rank.cell(row=row, column=1, value=metro['rank'])
        ws_rank.cell(row=row, column=2, value=metro['metro_name'])
        ws_rank.cell(row=row, column=3, value=metro['primary_city'])
        ws_rank.cell(row=row, column=4, value=metro['weighted_score'])
        ws_rank.cell(row=row, column=5, value=metro['grade']['letter'])
        ws_rank.cell(row=row, column=6, value=metro['grade']['percentile'])
        ws_rank.cell(row=row, column=7, value=metro['raw_values']['101A_unemployment'])
        ws_rank.cell(row=row, column=8, value=metro['raw_values']['102A_clf_yoy'])
        ws_rank.cell(row=row, column=9, value=metro['raw_values']['103B_earnings_yoy'])
        ws_rank.cell(row=row, column=10, value=metro['raw_values']['104C_col'])
        
        grade_color = grade_colors.get(metro['grade']['letter'], 'FFFFFF')
        for col in range(1, len(headers_rank) + 1):
            cell = ws_rank.cell(row=row, column=col)
            style_cell(cell, fill_color=grade_color)
            if col in [1, 4, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal='center')
    
    ws_rank.column_dimensions['A'].width = 8
    ws_rank.column_dimensions['B'].width = 35
    ws_rank.column_dimensions['C'].width = 18
    for col in range(4, 11):
        ws_rank.column_dimensions[get_column_letter(col)].width = 14
    
    # ========== SHEET 5: TOP VS BOTTOM ==========
    ws_compare = wb.create_sheet("Top vs Bottom", 4)
    
    ws_compare.cell(row=1, column=1, value="TOP 10 METROS")
    top_header = ws_compare.cell(row=1, column=1)
    top_header.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    top_header.font = Font(bold=True, color="FFFFFF", size=12)
    
    ws_compare.cell(row=1, column=8, value="BOTTOM 10 METROS")
    bottom_header = ws_compare.cell(row=1, column=8)
    bottom_header.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    bottom_header.font = Font(bold=True, color="FFFFFF", size=12)
    
    headers_compare = ['#', 'Metro', 'Score', 'Grade', '#', 'Metro', 'Score', 'Grade']
    for col, header in enumerate(headers_compare[:4], 1):
        cell = ws_compare.cell(row=2, column=col, value=header)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        style_cell(cell, bold=True, center=True)
    
    for col, header in enumerate(headers_compare[4:], 5):
        cell = ws_compare.cell(row=2, column=col, value=header)
        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        style_cell(cell, bold=True, center=True)
    
    for idx, metro in enumerate(sorted_metros[:10], 1):
        row = idx + 2
        ws_compare.cell(row=row, column=1, value=idx)
        ws_compare.cell(row=row, column=2, value=metro['metro_name'])
        ws_compare.cell(row=row, column=3, value=metro['weighted_score'])
        ws_compare.cell(row=row, column=4, value=metro['grade']['letter'])
        
        for col in range(1, 5):
            cell = ws_compare.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            style_cell(cell, center=True)
    
    for idx, metro in enumerate(sorted_metros[-10:], 1):
        row = idx + 2
        ws_compare.cell(row=row, column=5, value=idx)
        ws_compare.cell(row=row, column=6, value=metro['metro_name'])
        ws_compare.cell(row=row, column=7, value=metro['weighted_score'])
        ws_compare.cell(row=row, column=8, value=metro['grade']['letter'])
        
        for col in range(5, 9):
            cell = ws_compare.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            style_cell(cell, center=True)
    
    for col in range(1, 9):
        ws_compare.column_dimensions[get_column_letter(col)].width = 18
    
    # Save workbook
    excel_path = SCRIPT_DIR / 'Economic_Metrics_All_Metros.xlsx'
    wb.save(excel_path)
    print(f"✓ Excel file created: {excel_path}")
    return excel_path


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main entry point"""
    output = calculate_metrics()
    
    if not output:
        print("\n❌ Failed to calculate metrics")
        return
    
    # Save JSON output
    json_path = SCRIPT_DIR / 'calculated_metrics_reconciled.json'
    
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2)
        print(f"\n✓ JSON saved to: {json_path}")
    except Exception as e:
        print(f"\n❌ ERROR saving JSON: {e}")
        return
    
    # Create Excel file (YOUR RECEIPT!)
    try:
        excel_path = create_excel_from_metrics(output)
    except Exception as e:
        print(f"\n❌ ERROR creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Show summary
    print("\n" + "=" * 80)
    print("SUMMARY - PERCENTILE-BASED SCORING WITH 3-COMPONENT COL & 2-COMPONENT OWR")
    print("=" * 80)
    
    sorted_metros = sorted(output['metros'], key=lambda x: x['weighted_score'], reverse=True)
    
    print("\nTOP 10 METROS (by percentile score):")
    for i, metro in enumerate(sorted_metros[:10], 1):
        col_score = metro['raw_values']['104C_col']
        owr_score = metro['raw_values']['105C_owr']
        percentile = metro['weighted_percentile']
        col_str = f"{col_score:.2f}" if col_score is not None else "N/A"
        owr_str = f"{owr_score:.2f}" if owr_score is not None else "N/A"
        print(f"  {i:2d}. {metro['metro_name']:<40} {metro['grade_display']:8s} ({percentile:5.1f}th percentile) | COL: {col_str} OWR: {owr_str}")

    print("\nBOTTOM 5 METROS:")
    for i, metro in enumerate(sorted_metros[-5:], 1):
        col_score = metro['raw_values']['104C_col']
        owr_score = metro['raw_values']['105C_owr']
        percentile = metro['weighted_percentile']
        col_str = f"{col_score:.2f}" if col_score is not None else "N/A"
        owr_str = f"{owr_score:.2f}" if owr_score is not None else "N/A"
        print(f"  {i}. {metro['metro_name']:<40} {metro['grade_display']:8s} ({percentile:5.1f}th percentile) | COL: {col_str} OWR: {owr_str}")
    
    print("\n" + "=" * 80)
    print("📊 OUTPUT FILES GENERATED")
    print("=" * 80)
    print(f"✓ JSON Data ......... calculated_metrics_reconciled.json")
    print(f"✓ Excel Receipt ..... Economic_Metrics_All_Metros.xlsx")
    print("\n📈 Your Excel Receipt includes:")
    print("   • Sheet 1: Summary (all metros sorted by score)")
    print("   • Sheet 2: Metric Scores (individual 0-100 scores)")
    print("   • Sheet 3: Raw Values (actual metrics for verification)")
    print("   • Sheet 4: By Rank (original MSA ranking)")
    print("   • Sheet 5: Top vs Bottom (leaders vs laggards)")
    print("\n✅ READY FOR LINKEDIN CONTENT GENERATION\n")
    print("📊 Current Scoring — 8 Metrics (April 2026):")
    print("  Weights: 107E LDC 25% | 101A Unemp 20% | 103B Earnings 15% | 104C COL 12%")
    print("           102A LFP 10% | 200B Permits 10% | 204A DoM 5% | 105C OWR 3%")
    print("  Split: 85% Employment / 15% Housing\n")
    print("  Key design decisions:")
    print("  • 107E Labor Demand Composite: employment growth (70%) + hours deviation (30%)")
    print("    - Hours signal conditioned on employment direction — fixes survivor squeeze")
    print("    - Absorbs former standalone 106D (10%) + 107E (15%) = 25% combined")
    print("  • 204A DoM: 60-day inflection — rising DoM penalized above 60 days (demand destruction)")
    print("    - Below 60 days: loosening = good (inventory for workers)")
    print("    - Above 60 days: loosening = bad (owners locked in, labor mobility impaired)")
    print("  • 102A LFP reduced 15%→10%: annual population benchmark anchor limits dynamic value")
    print("  • 103B Earnings increased 10%→15%: real-time monthly signal, redistributed from LFP")
    print("  • 104C COL increased 10%→12%: redistributed from OWR reduction")
    print("  • 105C OWR reduced 5%→3%: tiebreaker only, avoids penalising industrial/energy metros\n")


if __name__ == '__main__':
    main()
